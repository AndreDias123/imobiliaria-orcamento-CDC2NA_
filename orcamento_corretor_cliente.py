from __future__ import annotations
from dataclasses import dataclass
from abc import ABC, abstractmethod
from typing import Optional, List, Dict
import sys, os, platform

# Excel deps
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instale a dependência primeiro:  pip install openpyxl")
    sys.exit(1)

CONTRACT_FEE = 2000.00
MAX_INSTALLMENTS = 5

# ----------------- Modelos -----------------
class Imovel(ABC):
    @abstractmethod
    def calcular_aluguel(self) -> float: ...

from dataclasses import dataclass

@dataclass
class Apartamento(Imovel):
    quartos: int = 1
    vagas: int = 0
    tem_criancas: bool = True
    def calcular_aluguel(self) -> float:
        v = 700.0
        if self.quartos == 2: v += 200.0
        if self.vagas > 0: v += 300.0
        if not self.tem_criancas: v *= 0.95
        return round(v, 2)

@dataclass
class Casa(Imovel):
    quartos: int = 1
    vagas: int = 0
    def calcular_aluguel(self) -> float:
        v = 900.0
        if self.quartos == 2: v += 250.0
        if self.vagas > 0: v += 300.0
        return round(v, 2)

@dataclass
class Estudio(Imovel):
    vagas: int = 0
    def calcular_aluguel(self) -> float:
        v = 1200.0
        if self.vagas == 0: return round(v, 2)
        if self.vagas <= 2: return round(v + 250.0, 2)
        return round(v + 250.0 + (self.vagas-2)*60.0, 2)

# ----------------- Serviços -----------------
@dataclass
class Orcamento:
    imovel: Imovel
    parcelas_contrato: int
    def validar(self) -> None:
        if not (1 <= self.parcelas_contrato <= MAX_INSTALLMENTS):
            raise ValueError(f"Parcelas do contrato devem estar entre 1 e {MAX_INSTALLMENTS}.")
    def aluguel_mensal(self) -> float: return round(self.imovel.calcular_aluguel(), 2)
    def parcela_contrato(self) -> float: return round(CONTRACT_FEE / self.parcelas_contrato, 2)
    def total_mensal(self) -> float: return round(self.aluguel_mensal() + self.parcela_contrato(), 2)
    def gerar_parcelas_12_meses(self) -> List[Dict[str, float]]:
        aluguel = self.aluguel_mensal(); parcela = self.parcela_contrato()
        out = []
        for mes in range(1, 13):
            pc = parcela if mes <= self.parcelas_contrato else 0.0
            out.append({"mes": mes, "aluguel": aluguel, "parcela_contrato": pc, "total": round(aluguel+pc,2)})
        return out

# ----------------- Excel -----------------
def salvar_excel_formatado(orc: Orcamento, caminho: str, corretor: str, cliente: str) -> None:
    dados = orc.gerar_parcelas_12_meses()
    wb = Workbook(); ws = wb.active; ws.title = "Orçamento"

    # estilos
    title_font = Font(name="Calibri", size=16, bold=True)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    header_fill = PatternFill("solid", fgColor="E6F2FF")
    zebra_fill = PatternFill("solid", fgColor="F9F9F9")
    border_thin = Border(*(Side(style="thin", color="CCCCCC") for _ in range(4)))
    moeda = NamedStyle(name="moeda"); moeda.number_format = '"R$" #,##0.00'
    if "moeda" not in wb.named_styles: wb.add_named_style(moeda)

    # cabeçalho/título
    ws.merge_cells("A1:D1"); ws["A1"] = "Orçamento de Aluguel – R.M Imobiliária"
    ws["A1"].font = title_font; ws["A1"].alignment = center
    ws["A2"] = f"Corretor: {corretor}"
    ws["A3"] = f"Cliente: {cliente}"

    # resumo à direita
    ws["F1"] = "Resumo"; ws["F1"].font = title_font
    ws["F3"] = "Aluguel mensal"; ws["G3"] = orc.aluguel_mensal(); ws["G3"].style="moeda"; ws["G3"].alignment=right
    ws["F4"] = "Parcela do contrato"; ws["G4"] = orc.parcela_contrato(); ws["G4"].style="moeda"; ws["G4"].alignment=right
    ws["F5"] = "Total mensal"; ws["G5"] = orc.total_mensal(); ws["G5"].style="moeda"; ws["G5"].alignment=right

    # cabeçalhos da tabela (começa na linha 5)
    headers = ["Mês", "Aluguel", "Parcela do contrato", "Total"]
    header_row = 5
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font=header_font; c.alignment=center; c.fill=header_fill; c.border=border_thin
    ws.freeze_panes = "A6"

    # dados
    row = header_row + 1
    for i, linha in enumerate(dados):
        ws.cell(row=row, column=1, value=linha["mes"]).alignment=center; ws.cell(row=row, column=1).border=border_thin
        ws.cell(row=row, column=2, value=linha["aluguel"]).style="moeda"; ws.cell(row=row, column=2).alignment=right; ws.cell(row=row, column=2).border=border_thin
        ws.cell(row=row, column=3, value=linha["parcela_contrato"]).style="moeda"; ws.cell(row=row, column=3).alignment=right; ws.cell(row=row, column=3).border=border_thin
        ws.cell(row=row, column=4, value=linha["total"]).style="moeda"; ws.cell(row=row, column=4).alignment=right; ws.cell(row=row, column=4).border=border_thin
        if i % 2 == 0:
            for c in range(1,5): ws.cell(row=row, column=c).fill = zebra_fill
        row += 1

    # larguras
    for col in range(1,5):
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row+1))
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(28, max_len+2))
    ws.column_dimensions["F"].width = 22; ws.column_dimensions["G"].width = 16

    wb.save(caminho)

def abrir_arquivo(caminho: str) -> None:
    try:
        s = platform.system()
        if s == "Windows": os.startfile(caminho)  # type: ignore[attr-defined]
        elif s == "Darwin": os.system(f'open "{caminho}"')
        else: os.system(f'xdg-open "{caminho}"')
    except Exception as e:
        print("Não foi possível abrir automaticamente. Caminho:", caminho, "| Erro:", e)

# ----------------- CLI -----------------
def ler_int(msg, minimo=None, maximo=None):
    while True:
        try:
            v = int(input(msg).strip())
            if minimo is not None and v < minimo: print(f"≥ {minimo}"); continue
            if maximo is not None and v > maximo: print(f"≤ {maximo}"); continue
            return v
        except ValueError: print("Digite um número inteiro válido.")

def ler_bool(msg):
    while True:
        v = input(msg + " [s/n]: ").strip().lower()
        if v in {"s","sim"}: return True
        if v in {"n","nao","não"}: return False
        print("Responda com 's' ou 'n'.")

def main():
    print("\n=== Orçamento - Excel ===\n")
    corretor = input("Nome do corretor: ").strip() or "—"
    cliente  = input("Nome do cliente: ").strip() or "—"

    print("\nTipo de imóvel:\n  1) Apartamento\n  2) Casa\n  3) Estúdio")
    tipo = input("Escolha [1-3]: ").strip()
    if tipo == "1":
        quartos = ler_int("Número de quartos (1 ou 2): ", 1, 2)
        vagas = ler_int("Vagas de garagem (0+): ", 0, None)
        tem_criancas = ler_bool("Há crianças no domicílio?")
        imovel = Apartamento(quartos, vagas, tem_criancas)
    elif tipo == "2":
        quartos = ler_int("Número de quartos (1 ou 2): ", 1, 2)
        vagas = ler_int("Vagas de garagem (0+): ", 0, None)
        imovel = Casa(quartos, vagas)
    else:
        vagas = ler_int("Vagas de estacionamento (0+): ", 0, None)
        imovel = Estudio(vagas)

    parcelas = ler_int(f"Nº de parcelas do contrato [1..{MAX_INSTALLMENTS}]: ", 1, MAX_INSTALLMENTS)

    orc = Orcamento(imovel=imovel, parcelas_contrato=parcelas); orc.validar()
    print("\n--- Resultado ---")
    print("Aluguel mensal: R$", orc.aluguel_mensal())
    print("Parcela do contrato: R$", orc.parcela_contrato())
    print("Total mensal (com parcela): R$", orc.total_mensal())

    nome = input("\nNome do arquivo Excel (ENTER para 'parcelas_orcamento.xlsx'): ").strip() or "parcelas_orcamento.xlsx"
    salvar_excel_formatado(orc, nome, corretor, cliente)
    print("Arquivo Excel salvo em:", nome)
    abrir_arquivo(nome)

if __name__ == "__main__":
    main()
